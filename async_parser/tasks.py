from celery import shared_task
from .query import WriteQuery
from .parser import ParseXml, ParseExcel, ParseApi
from provernalog.models import *
import openpyxl
from async_parser.models import TaskFile
from django.db import connection
import datetime


class TaskFlow:

    def check_group_appraise(self):
        groups = GroupAppraise.objects.filter(region=self.region, type=self.group_type)
        if not groups.exists():
            groups = GroupAppraise.objects.filter(region=78, type=self.group_type)
            for group in groups:
                GroupAppraise.objects.create(group_id=group.group_id, name=group.name, type=type, region=self.region)
        else:
            print("Группы оценки с заданными параметрами уже существуют")

    def create_partion_parcel_table(self):
        with connection.cursor() as cursor:
            cursor.execute("CREATE TABLE if not exists parcel_%s (LIKE parcel including all)", [self.region])
            cursor.execute("ALTER TABLE parcel_%s add constraint check_region check (region=%s)", [self.region, self.region])
            cursor.execute("ALTER TABLE parcel_%s inherit parcel", [self.region])


class ParcelTaskFlow(TaskFlow):
    def __init__(self, file, task_file_id, region, group_type, date_relevance, excel_name=None):
        self.file = file
        self.task_file = TaskFile.objects.get(pk=task_file_id)
        self.region = region
        self.group_type = group_type
        if self.group_type:
            self.check_group_appraise()
        if not excel_name:
            self.excel_name = {'cadastral_number': 0, 'group_appraise': 1}
        else:
            self.excel_name = excel_name
        self.date_relevance = date_relevance
        self.query = WriteQuery(region=self.region, task_file=self.task_file,
                                group_type=self.group_type, date_relevance=self.date_relevance)

    def parse(self):
        _prefix = self.file.split('.')[-1]
        if _prefix == 'xlsx':
            parse = ParseExcel(self.file, name=self.excel_name)
            parcels = parse.write_excel_parcel()
            factors = None
        elif _prefix == 'xml':
            parse = ParseXml(self.file)
            self.task_file.parcels = parse.get_count_parcels()
            self.task_file.save()
            factors = parse.get_factors()
            parcels = parse.get_parcel()
        for parcel in parcels:
            self.query.write_parcel(parcel, factors)
        self.task_file.is_parsed = True
        self.task_file.date_parsed = datetime.datetime.now()
        self.task_file.save()


@shared_task
def add_parcel(file, task_file_id, region, group_type, date_relevance):
    parse = ParcelTaskFlow(file, task_file_id, region, group_type, date_relevance)
    parse.parse()


@shared_task(bind=True)
def fetch_rosreestr(self, parcels):
    parser = ParseApi()
    for index, parcel in enumerate(parcels):
        if not self.request.called_directly:
            self.update_state(state='PROGRESS',
                              meta={'current': index, 'total': len(parcels)})
        parcel_query = Parcel.objects.get(cadastral_number=parcel["cadastral_number"])
        parcel = parser.low_priority(parcel_query.cadastral_number)
        if parcel:
            if parcel.get('current_cost'):
                parcel_query.current_cost = parcel['current_cost']
                parcel_query.save()


def convert_excel_email(filter):
    parcels = Parcel.objects.filter(**filter).values("cadastral_number", "area",
                                                     "address", "current_cost",
                                                     "cost_intermediate", "approved_cost")
    wb = openpyxl.Workbook()
    ws = wb["Sheet"]
    ws.column_dimensions["A"].width = 25.0
    ws.column_dimensions["D"].width = 20.0
    ws.column_dimensions["E"].width = 25.0
    ws.column_dimensions["F"].width = 25.0
    for value, item in {"A1": "Кадастровый номер",
              'B1': "Площадь", "C1": "Адрес",
              "D1": "Текущая стоимость",
              "E1": "Промежуточная стоимость",
              "F1": "Утвержденная стоимость"}.items():
        ws[value] = item
    for index, parcel in enumerate(parcels, start=2):
        ws[f"A{index}"] = parcel["cadastral_number"]
        ws[f"B{index}"] = parcel["area"]
        ws[f"C{index}"] = parcel["address"]
        ws[f"D{index}"] = parcel["current_cost"]
        ws[f"E{index}"] = parcel["cost_intermediate"]
        ws[f"F{index}"] = parcel["approved_cost"]
    return wb
