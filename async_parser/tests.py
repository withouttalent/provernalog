from django.test import TestCase
import os
from async_parser.utils import *
from django.forms.models import model_to_dict
from async_parser.query import WriteQuery
from async_parser.parser import ParseXml, ParseApi, ParseExcel
from provernalog.models import Parcel, ValueFactor, Utilization, GroupAppraise
import openpyxl
from openpyxl.writer.excel import save_virtual_workbook

class ParserTest(TestCase):

    def fill_file(self, text):
        file = open("check.xml", "w")
        file.write(text)
        file.close()
        return file

    def test_parser_fd_factor_xml(self):
        file = self.fill_file(text=text_fd_factor_state)
        p = ParseXml(file.name)
        factors = p.get_factors()
        check_sum = [{'id_factor': '3', 'name': 'Местоположение', 'description': 'Фактор отражающий изменение цены 1 кв.м. при изменении площади объекта недвижимости', 'quantitative_dimension': '', 'id': ['3_1', '3_2', '3_3', '3_4', '3_5'], 'value': ['I', ' II', ' III', ' IV', ' V']}]
        os.remove("check.xml")
        self.assertEqual(check_sum, factors)


    def test_parser_fd_parcel_xml(self):
        file = self.fill_file(text_fd_parcel_state)
        p = ParseXml(file.name)
        parcels = p.get_parcel()
        check_sum = [{'evaluative_factors': [], 'values': [], 'cadastral_number': '51:02:0060301:34', 'area': '20061'}]
        os.remove("check.xml")
        self.assertEqual(check_sum, parcels)

    def test_excel(self):
        name = {"cadastral_number": 0, "approved_cost": 1}
        wb = openpyxl.Workbook()
        ws = wb["Sheet"]
        ws["A1"] = "CN"
        ws["B1"] = "AC"
        ws["A2"] = "00:00:00:05"
        ws["B2"] = "1423.00"
        wb.save("check.xlsx")
        p = ParseExcel(file="check.xlsx", name=name)
        parcels = p.write_excel_parcel()
        os.remove("check.xlsx")
        self.assertEqual([{"cadastral_number": "00:00:00:05", "approved_cost": "1423.00"}], parcels)





class QueryTest(TestCase):


    def setUp(self):
        Utilization.objects.create(code=141001000000, name="test")
        GroupAppraise.objects.create(group_id=4, type="ZU", name="test", region=68)

    def test_query_parcel(self):
        parcel = {"cadastral_number": "00:00:00000:04",
                   "area": 20061.0, "current_cost": "4000.63",
                  "bydoc": "common information test юникод символ",
                  "assignation": "просто текст", "group_appraise": "4",
                  "utilization": "141001000000",
                  'address': 'Пензенская обл., р-н Белинский, с. Крюково, ул. Центральная, д. 7',
                  'evaluative_factors': ['0', '1', '2', '4'],
                  'values': ['105.5', '15', '320', '20'],
                  }
        factors = [
            {'id_factor': '0', 'name': 'Расстояние от  населенного пункта до столицы субъекта РФ',
             'description': 'Расстояние от  населенного пункта до столицы субъекта РФ', 'quantitative_dimension': 'км',
             'id': [], 'value': []},
            {'id_factor': '1',
             'name': 'Расстояние от населенного пункта до центра муниципального района, городского округа',
             'description': 'Расстояние от населенного пункта до центра муниципального района, городского округа',
             'quantitative_dimension': 'км', 'id': [], 'value': []},
            {'id_factor': '2', 'name': 'Численность населенного пункта',
             'description': 'Численность населенного пункта', 'quantitative_dimension': 'чел', 'id': [], 'value': []},
            {'id_factor': '4', 'name': 'Наличие в населенном пункте магазина',
             'description': 'Наличие в населенном пункте магазина', 'quantitative_dimension': '',
             'id': ['12', '13', '14', '15', '16', '17', '18', '19', '20', '21'],
             'value': ['Нет ', '        да', '       да', ' да', '    нет', ' нет', 'Да', 'Нет', 'да', 'нет']},
        ]
        w = WriteQuery(region=68, group_type="ZU", date_relevance="2018-07-21")
        w.write_parcel(parcels=[parcel.copy()], factors=factors)
        parcel_query = Parcel.objects.get(cadastral_number="00:00:00000:04")
        _parcel = {"cadastral_number": parcel_query.cadastral_number, "area": parcel_query.area,
                        "current_cost": str(parcel_query.current_cost), "bydoc": parcel_query.bydoc,
                        "assignation": parcel_query.assignation,
                        "group_appraise": str(parcel_query.group_appraise.group_id),
                        "utilization": parcel_query.utilization.code,
                        "address": parcel_query.address,
                        }
        del parcel["evaluative_factors"], parcel["values"]
        self.assertEqual(parcel, _parcel)
        value_factor = ValueFactor.objects.filter(parcel=parcel_query)
        self.assertTrue(value_factor.exists())
        self.assertEqual(value_factor.count(), 4)

